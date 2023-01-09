from openpyxl import Workbook
# dictionary
week_temperatures = {
    'Monday': 54,
    'Tuesday': 60,
    'Wednesday': 62,
    'Thursday': 57,
    'Friday': 71
}
# creating workbook
workbook = Workbook()
worksheet = workbook.active
worksheet.title = 'Daily Temperatures'

worksheet.cell(1, 1, 'Day')
worksheet.cell(1, 2, 'Temperatures (F)')

for index, day in enumerate(week_temperatures):
    row = index + 2
    temp = week_temperatures[day]
    worksheet.cell(row, 1, day)
    worksheet.cell(row, 2, temp)

workbook.save('temperatures.xlsx')