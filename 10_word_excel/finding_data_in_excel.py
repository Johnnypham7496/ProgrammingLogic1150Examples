# openpyxl will help read Excel file and will have to install openpyxl interpreter
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
# this selects the workbook within the same folder and attaches it to a variable
wb = openpyxl.load_workbook('ITEC_Courses.xlsx')
print(type(wb))
print(wb.sheetnames)
print()

# shows the rooms' worksheet, tells us the type, tells us the title
sheet_2 = wb['rooms']
print(type(sheet_2))
print(sheet_2.title)
print()

# indicates another sheet that is active which in considered as sheet 1
another_sheet = wb.active
print(another_sheet)
print()

sheet_2_cell_data = sheet_2['A1']
sheet_2_cell_data = sheet_2_cell_data.value
print(sheet_2_cell_data)
# for i in range(0, 7, 2):
#    print(i, sheet_2_cell_data.cell(row = 1, column= 2).value)
print()

print(get_column_letter(3))
print()

for row_of_cell_objects in sheet_2_cell_data['A1':'C3']:
    for cell_obj in row_of_cell_objects:
        print(cell_obj.coordinate, cell_obj.value)
    print('--- END OF ROW ---')
